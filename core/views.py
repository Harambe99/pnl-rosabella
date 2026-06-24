"""Views: login, dashboard (Monthly + Daily P&L), uploads, Monthly Inputs, COGS, history."""
from datetime import date
from calendar import monthrange
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt

from django.db.models import Q
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from .models import (COGSItem, MonthlyInput, ImportLog, MonthlyInputAudit,
                     AdLedgerDay, AdLedgerConfig, AgencyPromoTag, AgencyInvoice,
                     AdTransaction, Order, SettlementRow, SellerShipmentCost,
                     AnalyticsDay, AdSpendDay)
from .aggregator import compute_daily_pnl, compute_monthly_pnl, PNL_ROW_LAYOUT
from .importers import (import_manage_orders, import_settlement,
                        import_shop_analytics, import_ad_spend, import_fbt_billing,
                        import_seller_shipping, import_ad_transactions)


def login_view(request):
    if request.method == 'POST':
        pwd = request.POST.get('password', '')
        if pwd == settings.APP_PASSWORD:
            request.session['app_authed'] = True
            request.session.set_expiry(60 * 60 * 24 * 30)
            return redirect(request.GET.get('next') or 'dashboard')
        messages.error(request, 'Wrong password.')
    return render(request, 'core/login.html')


def logout_view(request):
    request.session.flush()
    return redirect('login')


def _pct(v, nr):
    """Return value/net_revenue as a percentage, or None if not meaningful."""
    if v is None or nr is None: return None
    try:
        nr_f = float(nr)
        if nr_f == 0: return None
        return float(v) / nr_f * 100
    except (TypeError, ValueError):
        return None


def dashboard(request):
    try:
        year = int(request.GET.get('year', 2026))
        # Sanity-clamp to a sensible range so date() construction doesn't blow up
        if year < 2020 or year > 2099:
            return redirect('dashboard')
    except (TypeError, ValueError):
        return redirect('dashboard')
    monthly = compute_monthly_pnl(year)
    months = [f'{year}-{m:02d}' for m in range(1, 13)]
    rows = []
    # Pre-compute Net Revenue per month (for % column) and YTD Net Revenue
    nr_per_month = [monthly.get(m, {}).get('NET REVENUE') for m in months]
    ytd_nr = sum((nr or Decimal('0')) for nr in nr_per_month)
    for label, rtype in PNL_ROW_LAYOUT:
        if rtype == 'blank':
            rows.append({'label': '', 'type': 'blank', 'cells': [(None, None)]*12,
                         'ytd': None, 'ytd_pct': None})
            continue
        if rtype in ('section', 'sub'):
            rows.append({'label': label, 'type': rtype, 'cells': [(None, None)]*12,
                         'ytd': None, 'ytd_pct': None})
            continue
        cells = []
        ytd = Decimal('0')
        for m, nr in zip(months, nr_per_month):
            v = monthly.get(m, {}).get(label)
            cells.append((v, _pct(v, nr)))
            if v is not None: ytd += v
        ytd_pct = _pct(ytd, ytd_nr) if ytd_nr else None
        rows.append({'label': label, 'type': rtype, 'cells': cells,
                     'ytd': ytd, 'ytd_pct': ytd_pct})
    return render(request, 'core/dashboard.html', {
        'rows': rows, 'months': months, 'year': year,
    })


def daily_view(request):
    """Render the FULL year of daily P&L in one scrollable table.
    The month picker is a 'scroll-to' anchor, not a filter."""
    yyyy_mm = request.GET.get('month') or date.today().strftime('%Y-%m')
    try:
        y, m = int(yyyy_mm[:4]), int(yyyy_mm[5:7])
        if y < 2020 or y > 2099 or m < 1 or m > 12:
            return redirect('daily')
    except Exception:
        return redirect('daily')
    start = date(y, 1, 1)
    end = date(y, 12, 31)
    daily = compute_daily_pnl(start, end)
    dates = sorted(daily.keys())
    nr_per_date = [daily.get(d, {}).get('NET REVENUE') for d in dates]
    rows = []
    for label, rtype in PNL_ROW_LAYOUT:
        if rtype == 'blank':
            rows.append({'label': '', 'type': 'blank', 'cells': [(None, None)]*len(dates)})
            continue
        if rtype in ('section', 'sub'):
            rows.append({'label': label, 'type': rtype, 'cells': [(None, None)]*len(dates)})
            continue
        cells = []
        for d, nr in zip(dates, nr_per_date):
            v = daily.get(d, {}).get(label)
            cells.append((v, _pct(v, nr)))
        rows.append({'label': label, 'type': rtype, 'cells': cells})
    return render(request, 'core/daily.html', {
        'rows': rows, 'dates': dates, 'yyyy_mm': yyyy_mm, 'year': y,
        'prev_year': y - 1, 'next_year': y + 1,
    })


def upload(request):
    if request.method == 'POST':
        kind = request.POST.get('kind')
        f = request.FILES.get('file')
        if not f or not kind:
            messages.error(request, 'Missing file or import type.')
            return redirect('upload')
        try:
            if kind == 'manage_orders':
                result = import_manage_orders(f, f.name)
            elif kind == 'settlement':
                result = import_settlement(f, f.name)
            elif kind == 'shop_analytics':
                result = import_shop_analytics(f, f.name)
            elif kind == 'ad_spend':
                result = import_ad_spend(f, f.name)
            elif kind == 'fbt_billing':
                period = request.POST.get('period', '').strip()
                result = import_fbt_billing(f, period, f.name)
            elif kind == 'seller_shipping':
                result = import_seller_shipping(f, f.name)
            elif kind == 'ad_transactions':
                result = import_ad_transactions(f, f.name)
            else:
                messages.error(request, 'Unknown import type.')
                return redirect('upload')
            messages.success(request, f'Import done: {result}')
        except Exception as e:
            messages.error(request, f'Import error: {e}')
        return redirect('upload')
    recent = ImportLog.objects.all()[:20]
    return render(request, 'core/upload.html', {'recent': recent})


def cogs(request):
    if request.method == 'POST':
        for k, v in request.POST.items():
            if k.startswith('cogs_'):
                try:
                    pk = int(k.split('_')[1])
                    item = COGSItem.objects.get(pk=pk)
                    item.cogs_per_order = Decimal(v or '0')
                    item.save(update_fields=['cogs_per_order'])
                except Exception:
                    pass
        messages.success(request, 'COGS updated.')
        return redirect('cogs')
    items = COGSItem.objects.all()
    return render(request, 'core/cogs.html', {'items': items})


def monthly_inputs(request):
    if request.method == 'POST':
        month = request.POST.get('month')
        if month:
            mi, _ = MonthlyInput.objects.get_or_create(month=month)
            changes = []
            for field in MonthlyInput._meta.get_fields():
                if field.name in ('id', 'month', 'updated_at'): continue
                if not hasattr(field, 'attname'): continue
                raw = request.POST.get(field.name)
                if raw is None or str(raw).strip() == '':
                    continue  # empty input = leave existing value alone
                try: new_val = Decimal(raw)
                except: continue
                old_val = getattr(mi, field.name) or Decimal('0')
                if new_val != old_val:
                    changes.append((field.name, old_val, new_val))
                    setattr(mi, field.name, new_val)
            mi.save()
            for fname, old, new in changes:
                MonthlyInputAudit.objects.create(month=month, field_name=fname,
                                                 old_value=old, new_value=new)
            if changes:
                messages.success(request, f'{month}: {len(changes)} field(s) updated.')
            else:
                messages.info(request, f'{month}: no changes (all fields empty or unchanged).')
        return redirect('monthly_inputs')
    months = MonthlyInput.objects.all()
    # Build JSON map {month: {field: value}} for the JS overwrite-warning
    import json as _json
    existing_data = {}
    for m in months:
        existing_data[m.month] = {}
        for field in MonthlyInput._meta.get_fields():
            if field.name in ('id', 'month', 'updated_at'): continue
            if not hasattr(field, 'attname'): continue
            existing_data[m.month][field.name] = str(getattr(m, field.name) or 0)
    audits = MonthlyInputAudit.objects.all()[:100]
    return render(request, 'core/monthly_inputs.html', {
        'months': months,
        'existing_json': _json.dumps(existing_data),
        'audits': audits,
    })


def history(request):
    logs = ImportLog.objects.all()[:200]
    return render(request, 'core/history.html', {'logs': logs})


def readme(request):
    return render(request, 'core/readme.html')


# ===========================================================================
# Export helpers — Source data sheets + working hyperlinks
# ===========================================================================

# Maps each P&L line label → the source sheet that backs it. Lines without
# an entry have no source (computed totals, section headers, manual inputs).
# Keys are stripped of leading whitespace to match LINE_ITEM_DOCS format —
# PNL_ROW_LAYOUT uses indented labels (e.g. '   FBT Fulfillment Fee') but
# lookups use the stripped form.
LINE_ITEM_TO_SOURCE = {
    'Gross Sales': 'Source — Manage Orders',
    'Less: Promos & Discounts': 'Source — Manage Orders',
    'COGS': 'Source — Manage Orders',
    'GMV (TikTok Analytics — reference)': 'Source — Shop Analytics',
    'Less: Refunds': 'Source — Settlement',
    'FBT Fulfillment Fee': 'Source — Settlement',
    'FBT Fulfillment Reimbursement': 'Source — Settlement',
    'TT Shop Shipping Incentive': 'Source — Settlement',
    'Shipping Fee Subsidy': 'Source — Settlement',
    'Customer Shipping Fee Offset': 'Source — Settlement',
    'Customer-Paid Shipping Fee': 'Source — Settlement',
    'Customer-Paid Shipping Refund': 'Source — Settlement',
    'Seller Shipping Fee Discount': 'Source — Settlement',
    'Logistics Reimbursement': 'Source — Settlement',
    'FBT Warehouse Compensation': 'Source — Settlement',
    'FBT Warehouse Service Fee': 'Source — Settlement',
    'Referral Fee': 'Source — Settlement',
    'Refund Admin Fee': 'Source — Settlement',
    'Campaign Service Fee': 'Source — Settlement',
    'Violation Fee': 'Source — Settlement',
    'TikTok Shop Reimb': 'Source — Settlement',
    'Rebate': 'Source — Settlement',
    'Co-funded Promotion (seller-funded)': 'Source — Settlement',
    'Co-funded Promotion Campaign Period Fee': 'Source — Settlement',
    'Chargebacks': 'Source — Settlement',
    'Unclassified Adjustments': 'Source — Settlement',
    'Platform (Affiliate Commission)': 'Source — Settlement',
    'Cost to Ship to FBT': 'Source — FBT Billing',
    'FBT Hub Placement Fee': 'Source — FBT Billing',
    'FBT Storage Fee': 'Source — FBT Billing',
    'FBT Inbound Shipping Fee': 'Source — FBT Billing',
    'FBT Inbound Incidents Fee': 'Source — FBT Billing',
    'FBT Booking Non-Compliance': 'Source — FBT Billing',
    'FBT Routing Non-Compliance': 'Source — FBT Billing',
    'FBT Outbound No-Show': 'Source — FBT Billing',
    'FBT Delayed Response Fee': 'Source — FBT Billing',
    'FBT Disposal Fee': 'Source — FBT Billing',
    'FBT Return Shipping (VAS)': 'Source — FBT Billing',
    'FBT Return to Seller Handling': 'Source — FBT Billing',
    'FBT Inbound Return Operation': 'Source — FBT Billing',
    'Cost to Ship to Customer': 'Source — Seller Shipping',
    'Ad Spend — Direct to TikTok (cash)': 'Source — Ad Spend',
    'Less: TBSM Savings': 'Source — Ad Ledger',
    'Less: TT Promo Credits': 'Source — Ad Ledger',
    # Manual monthly-input lines — no file source, but the Monthly Inputs sheet
    # lists every value with an empty Reference Link column for the user to paste
    # Google Sheet URLs explaining where each number came from.
    'Team Spend': 'Source — Monthly Inputs',
    'Software & Tools': 'Source — Monthly Inputs',
    'Other G&A': 'Source — Monthly Inputs',
    'Monthly Retainers': 'Source — Monthly Inputs',
    'Outsourced Agency': 'Source — Monthly Inputs',
    'Off-Platform (1% method)': 'Source — Monthly Inputs',
}


def _set_internal_link(cell, sheet_name, anchor='A1', link_font=None):
    """Set a proper internal hyperlink on a cell — works in Excel AND Google Sheets.
    openpyxl's string-style ``cell.hyperlink = "#'Sheet'!A1"`` writes XML that
    Google Sheets silently ignores. The explicit Hyperlink object writes the
    OOXML element both apps honor."""
    from openpyxl.worksheet.hyperlink import Hyperlink
    safe = sheet_name.replace("'", "''")
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{safe}'!{anchor}",
        display=str(cell.value) if cell.value is not None else '',
    )
    if link_font is not None:
        cell.font = link_font


def _build_source_sheets(wb, start_date, end_date, styles):
    """Build Source — * sheets covering [start_date, end_date]. Each sheet contains
    only the columns the importer actually reads. Sheets are grouped by source TYPE
    (Orders, Settlement, etc.) and rows are sorted chronologically so months in a
    multi-month export appear adjacent. Returns the set of sheet names actually
    created (some sources may have no rows in range and get skipped)."""
    F_HEAD = styles['head']; FILL_HEAD = styles['fill_head']
    F_TOTAL = styles['total']; FILL_TOTAL = styles['fill_total']
    F_MONTH = styles['month']; FILL_MONTH = styles['fill_month']
    BORDER = styles['border']
    DOLLAR = '$#,##0.00;[Red]($#,##0.00)'
    DATE_FMT = 'yyyy-mm-dd'

    created = set()
    months_in_range = []
    cy, cm = start_date.year, start_date.month
    while (cy, cm) <= (end_date.year, end_date.month):
        months_in_range.append((cy, cm))
        cm += 1
        if cm > 12: cm = 1; cy += 1

    # Memory safeguard — Render Standard tier OOMs at ~2GB.
    # The two heaviest source sheets (Manage Orders + Settlement) carry tens of
    # thousands of rows per month. For ranges >2 months we skip them to keep
    # the workbook fits in memory; the Guide tells users to export per-month
    # for full source data.
    BIG_SHEETS_ENABLED = len(months_in_range) <= 2
    skipped_for_size = set()  # passed back so the Guide can explain
    if not BIG_SHEETS_ENABLED:
        skipped_for_size.update({'Source — Manage Orders', 'Source — Settlement'})

    def _section_header(ws, row, label, n_cols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row, 1, label)
        c.font = F_MONTH; c.fill = FILL_MONTH
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20

    def _write_header(ws, headers, widths):
        for i, h in enumerate(headers, start=1):
            c = ws.cell(1, i, h)
            c.font = F_HEAD; c.fill = FILL_HEAD
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = BORDER
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28

    # --- Source — Manage Orders ---  (skipped for ranges >2 months — memory)
    # Mirrors aggregator COGS rule: Canceled orders are EXCLUDED from P&L COGS
    # unless they have a matching FBT fulfillment fee row (= shipped before
    # cancel). Pre-ship cancellations get a red COGS cell + 'No' in the new
    # 'Counted in P&L COGS?' column. P&L COGS subtotal ties exactly to the P&L.
    o_cols = ['Order Created Date', 'Order ID', 'SKU ID', 'Status', 'Quantity',
              'Gross Sales', 'Seller Discount (Promos)', 'COGS', 'Counted in P&L COGS?']
    o_widths = [16, 22, 22, 14, 9, 14, 16, 12, 18]
    o_rows = []
    shipped_canceled_ids = set()
    if BIG_SHEETS_ENABLED:
        o_qs = Order.objects.filter(
            created_date__gte=start_date, created_date__lte=end_date
        ).values_list('created_date', 'order_id', 'sku_id', 'status', 'quantity',
                      'gross_sale', 'seller_discount', 'cogs').order_by('created_date', 'order_id')
        o_rows = list(o_qs)
        # Order IDs that DID ship (have an FBT fulfillment fee row) — these
        # Canceled rows still count in P&L COGS. Anything Canceled NOT here
        # is a pre-ship cancellation and is excluded from P&L COGS.
        shipped_canceled_ids = set(SettlementRow.objects.filter(
            row_type='Order', fbt_fee__lt=0,
        ).values_list('order_id', flat=True))
    if o_rows:
        ws = wb.create_sheet('Source — Manage Orders')
        _write_header(ws, o_cols, o_widths)
        F_RED = Font(color='C0392B', size=11)
        F_RED_BOLD = Font(color='C0392B', size=11, bold=True)
        r = 2; current_month = None
        total_gross = total_disc = total_cogs = Decimal('0')
        total_cogs_counted = Decimal('0')
        total_cogs_excluded = Decimal('0')
        for od, oid, sku, status, qty, gross, disc, cogs in o_rows:
            mkey = (od.year, od.month)
            if mkey != current_month:
                _section_header(ws, r, od.strftime('%B %Y').upper(), len(o_cols))
                r += 1
                current_month = mkey
            is_canceled = (status or '').strip().lower() == 'canceled'
            counted = (not is_canceled) or (oid in shipped_canceled_ids)
            ws.cell(r, 1, od).number_format = DATE_FMT
            ws.cell(r, 2, oid); ws.cell(r, 3, sku); ws.cell(r, 4, status); ws.cell(r, 5, qty or 0)
            ws.cell(r, 6, float(gross or 0)).number_format = DOLLAR
            ws.cell(r, 7, float(disc or 0)).number_format = DOLLAR
            cogs_cell = ws.cell(r, 8, float(cogs or 0)); cogs_cell.number_format = DOLLAR
            if not counted:
                cogs_cell.font = F_RED  # visually flag: not deducted from P&L
            ws.cell(r, 9, 'Yes' if counted else 'No').alignment = Alignment(horizontal='center')
            total_gross += gross or Decimal('0')
            total_disc += disc or Decimal('0')
            total_cogs += cogs or Decimal('0')
            if counted:
                total_cogs_counted += cogs or Decimal('0')
            else:
                total_cogs_excluded += cogs or Decimal('0')
            r += 1
        # GRAND TOTAL — all orders (raw)
        ws.cell(r, 1, 'GRAND TOTAL (all orders)').font = F_TOTAL
        for col, val in zip([6, 7, 8], [total_gross, total_disc, total_cogs]):
            c = ws.cell(r, col, float(val)); c.number_format = DOLLAR
            c.font = F_TOTAL; c.fill = FILL_TOTAL
        r += 1
        # P&L COGS subtotal — excludes pre-ship cancellations (red rows above)
        ws.cell(r, 1, 'P&L COGS (excludes pre-ship cancellations)').font = F_TOTAL
        c = ws.cell(r, 8, float(total_cogs_counted)); c.number_format = DOLLAR
        c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.cell(r, 9, 'Sum of Yes').alignment = Alignment(horizontal='center')
        r += 1
        # Excluded COGS — pre-ship cancellations (visual context only)
        ws.cell(r, 1, 'Excluded — pre-ship cancellations').font = F_RED_BOLD
        c = ws.cell(r, 8, float(total_cogs_excluded)); c.number_format = DOLLAR
        c.font = F_RED_BOLD
        ws.cell(r, 9, 'Sum of No').alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        created.add('Source — Manage Orders')

    # --- Source — Settlement ---
    s_cols = ['Order Created Date', 'Statement Date', 'Order/Adjustment ID', 'Statement ID', 'Type', 'Qty',
              'Referral Fee', 'Refund Admin Fee', 'Campaign Service Fee',
              'Affiliate Total', 'FBT Fulfillment Fee', 'FBT Fulfillment Reimb',
              'Shipping (parent)', 'TT Shop Shipping Incentive', 'Shipping Fee Subsidy',
              'Customer Shipping Fee Offset', 'Customer-Paid Shipping Fee',
              'Customer-Paid Shipping Refund', 'Seller Shipping Fee Discount',
              'Co-funded Promo (seller)', 'Co-funded Promo Campaign Fee',
              'Refund Total', 'Chargeback', 'Violation Fee', 'TT Shop Reimb',
              'Logistics Reimb', 'FBT Warehouse Service Fee', 'FBT Warehouse Comp',
              'Rebate', 'Unclassified Adjustment']
    s_widths = [16, 14, 22, 18, 22, 6] + [14]*24
    s_rows = []
    if BIG_SHEETS_ENABLED:
        s_qs = SettlementRow.objects.filter(
            order_created_date__gte=start_date, order_created_date__lte=end_date
        ).values_list(
            'order_created_date', 'statement_date', 'order_id', 'settlement_id', 'row_type', 'quantity',
            'referral_fee', 'refund_admin', 'campaign_fee', 'affiliate_total',
            'fbt_fee', 'fbt_reimb', 'shipping', 'tt_shop_shipping_incentive',
            'shipping_fee_subsidy', 'customer_shipping_fee_offset', 'customer_paid_shipping_fee',
            'customer_paid_shipping_refund', 'seller_shipping_fee_discount',
            'cofunded_promo', 'cofunded_promo_campaign_fee', 'refund_total',
            'chargeback', 'violation', 'tt_shop_reimb', 'logistics_reimb',
            'fbt_warehouse', 'fbt_warehouse_comp', 'rebate', 'unclassified',
        ).order_by('order_created_date', 'order_id')
        s_rows = list(s_qs)
    if s_rows:
        ws = wb.create_sheet('Source — Settlement')
        _write_header(ws, s_cols, s_widths)
        r = 2; current_month = None
        n_money_cols = len(s_cols) - 6  # all cols after qty are money
        totals = [Decimal('0')] * n_money_cols
        for row_tuple in s_rows:
            od = row_tuple[0]
            mkey = (od.year, od.month) if od else None
            if mkey and mkey != current_month:
                _section_header(ws, r, od.strftime('%B %Y').upper(), len(s_cols))
                r += 1
                current_month = mkey
            for i, v in enumerate(row_tuple, start=1):
                if i == 1 or i == 2:  # dates
                    if v: ws.cell(r, i, v).number_format = DATE_FMT
                elif i <= 5:  # text/id
                    ws.cell(r, i, v if v is not None else '')
                elif i == 6:  # qty
                    ws.cell(r, i, v or 0)
                else:  # money
                    c = ws.cell(r, i, float(v or 0))
                    c.number_format = DOLLAR
                    totals[i - 7] += v or Decimal('0')
            r += 1
        ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
        for j, val in enumerate(totals):
            c = ws.cell(r, 7 + j, float(val)); c.number_format = DOLLAR
            c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.freeze_panes = 'A2'
        created.add('Source — Settlement')

    # --- Source — Seller Shipping ---
    sh_cols = ['Order Date', 'Shipped Date', 'Reference Number', 'Customer', 'Carrier',
               'Postage', 'Qty', 'Per Pack', 'Per Pick', 'Total']
    sh_widths = [12, 12, 26, 24, 18, 10, 6, 10, 10, 12]
    sh_qs = SellerShipmentCost.objects.filter(
        Q(order_date__gte=start_date, order_date__lte=end_date)
        | Q(order_date__isnull=True, shipped_date__gte=start_date, shipped_date__lte=end_date)
    ).order_by('order_date', 'shipped_date')
    sh_list = list(sh_qs.values_list(
        'order_date', 'shipped_date', 'reference_number', 'customer_name',
        'carrier_service', 'postage', 'product_quantity', 'per_pack', 'per_pick'))
    if sh_list:
        ws = wb.create_sheet('Source — Seller Shipping')
        _write_header(ws, sh_cols, sh_widths)
        r = 2; current_month = None
        t_postage = t_pack = t_pick = t_total = Decimal('0')
        for od, sd, ref, cust, carr, postage, qty, per_pack, per_pick in sh_list:
            anchor = od or sd
            mkey = (anchor.year, anchor.month) if anchor else None
            if mkey and mkey != current_month:
                _section_header(ws, r, anchor.strftime('%B %Y').upper(), len(sh_cols))
                r += 1
                current_month = mkey
            if od: ws.cell(r, 1, od).number_format = DATE_FMT
            if sd: ws.cell(r, 2, sd).number_format = DATE_FMT
            ws.cell(r, 3, ref or ''); ws.cell(r, 4, cust or ''); ws.cell(r, 5, carr or '')
            ws.cell(r, 6, float(postage or 0)).number_format = DOLLAR
            ws.cell(r, 7, qty or 0)
            ws.cell(r, 8, float(per_pack or 0)).number_format = DOLLAR
            ws.cell(r, 9, float(per_pick or 0)).number_format = DOLLAR
            total = (postage or Decimal('0')) + (per_pack or Decimal('0')) + (per_pick or Decimal('0'))
            ws.cell(r, 10, float(total)).number_format = DOLLAR
            t_postage += postage or Decimal('0'); t_pack += per_pack or Decimal('0')
            t_pick += per_pick or Decimal('0'); t_total += total
            r += 1
        ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
        for col, val in zip([6, 8, 9, 10], [t_postage, t_pack, t_pick, t_total]):
            c = ws.cell(r, col, float(val)); c.number_format = DOLLAR
            c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.freeze_panes = 'A2'
        created.add('Source — Seller Shipping')

    # --- Source — Monthly Inputs (manual entries: Team Spend, Software, etc.) ---
    # Rows = line items, columns = one per month + Total + Reference Link.
    # The Reference Link column is intentionally empty — the user pastes
    # Google Sheet URLs there after opening in Sheets to explain each value.
    mi_lines = [
        ('Team Spend',                'team_spend',         'Payroll, contractors — whoever you pay to operate the business.'),
        ('Software & Tools',          'software_tools',     'SaaS subscriptions (e.g. monthly tooling stack).'),
        ('Monthly Retainers',         'monthly_retainers',  'Creator retainers paid monthly.'),
        ('Outsourced Agency',         'creatify',           'Agency fees (DB field: creatify).'),
        ('Off-Platform (1% method)',  'off_platform_1pct',  'Estimated 1% commission for off-platform creators.'),
        ('Other G&A',                 'other_ga',           'Misc admin: legal, accounting, office, etc.'),
    ]
    months_set_inputs = {f'{y:04d}-{m:02d}' for y, m in months_in_range}
    mi_inputs_qs = MonthlyInput.objects.filter(month__in=months_set_inputs).order_by('month')
    mi_inputs_map = {mi.month: mi for mi in mi_inputs_qs}
    if mi_inputs_map:
        month_keys = sorted(months_set_inputs)
        mi_cols = ['Line Item', 'Description'] + month_keys + ['Total', 'Reference Link (paste Google Sheet URL)']
        mi_widths = [26, 60] + [14] * len(month_keys) + [14, 50]
        ws = wb.create_sheet('Source — Monthly Inputs')
        _write_header(ws, mi_cols, mi_widths)
        r = 2
        for label, field, desc in mi_lines:
            ws.cell(r, 1, label).font = F_TOTAL
            ws.cell(r, 2, desc).alignment = Alignment(wrap_text=True, vertical='top')
            line_total = Decimal('0')
            for j, mkey in enumerate(month_keys, start=3):
                mi = mi_inputs_map.get(mkey)
                v = (getattr(mi, field) if mi else None) or Decimal('0')
                c = ws.cell(r, j, float(v)); c.number_format = DOLLAR
                line_total += v
            c = ws.cell(r, 2 + len(month_keys) + 1, float(line_total))
            c.number_format = DOLLAR; c.font = F_TOTAL; c.fill = FILL_TOTAL
            # Reference Link column intentionally left blank for user to fill.
            ws.cell(r, 2 + len(month_keys) + 2, '').alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[r].height = 32
            r += 1
        # Column total row
        ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
        for j, mkey in enumerate(month_keys, start=3):
            col_total = Decimal('0')
            for _, field, _desc in mi_lines:
                mi = mi_inputs_map.get(mkey)
                col_total += (getattr(mi, field) if mi else None) or Decimal('0')
            c = ws.cell(r, j, float(col_total)); c.number_format = DOLLAR
            c.font = F_TOTAL; c.fill = FILL_TOTAL
        grand = Decimal('0')
        for _, field, _desc in mi_lines:
            for mi in mi_inputs_map.values():
                grand += (getattr(mi, field) or Decimal('0'))
        c = ws.cell(r, 2 + len(month_keys) + 1, float(grand))
        c.number_format = DOLLAR; c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.freeze_panes = 'C2'
        created.add('Source — Monthly Inputs')

    # --- Source — FBT Billing (one row per month) ---
    fbt_cols = ['Month', 'Cost to Ship to FBT', 'FBT Hub Placement Fee', 'FBT Storage Fee',
                'FBT Inbound Shipping Fee', 'FBT Inbound Incidents Fee',
                'FBT Booking Non-Compliance', 'FBT Routing Non-Compliance',
                'FBT Outbound No-Show', 'FBT Delayed Response Fee',
                'FBT Disposal Fee', 'FBT Return Shipping (VAS)',
                'FBT Return to Seller Handling', 'FBT Inbound Return Operation']
    fbt_widths = [11] + [16]*13
    fbt_fields = ['cost_ship_to_fbt', 'fbt_hub_placement', 'fbt_storage',
                  'fbt_inbound_shipping', 'fbt_inbound_incidents', 'fbt_booking_noncomp',
                  'fbt_routing_noncomp', 'fbt_outbound_noshow', 'fbt_delayed_response',
                  'fbt_disposal', 'fbt_return_shipping', 'fbt_return_seller_handling',
                  'fbt_inbound_return_op']
    months_set = {f'{y:04d}-{m:02d}' for y, m in months_in_range}
    mi_qs = MonthlyInput.objects.filter(month__in=months_set).order_by('month')
    mi_list = list(mi_qs)
    if mi_list:
        ws = wb.create_sheet('Source — FBT Billing')
        _write_header(ws, fbt_cols, fbt_widths)
        r = 2
        totals = [Decimal('0')] * len(fbt_fields)
        for mi in mi_list:
            ws.cell(r, 1, mi.month)
            for j, f in enumerate(fbt_fields):
                v = getattr(mi, f) or Decimal('0')
                c = ws.cell(r, 2 + j, float(v)); c.number_format = DOLLAR
                totals[j] += v
            r += 1
        ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
        for j, val in enumerate(totals):
            c = ws.cell(r, 2 + j, float(val)); c.number_format = DOLLAR
            c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.freeze_panes = 'A2'
        created.add('Source — FBT Billing')

    # --- Source — Shop Analytics ---
    an_qs = AnalyticsDay.objects.filter(date__gte=start_date, date__lte=end_date).order_by('date')
    an_list = list(an_qs.values_list('date', 'gmv', 'orders', 'items_sold'))
    if an_list:
        ws = wb.create_sheet('Source — Shop Analytics')
        _write_header(ws, ['Date', 'GMV', 'Orders', 'Items Sold'], [12, 14, 10, 10])
        r = 2; current_month = None
        t_gmv = Decimal('0'); t_orders = 0; t_items = 0
        for d, gmv, orders, items in an_list:
            mkey = (d.year, d.month)
            if mkey != current_month:
                _section_header(ws, r, d.strftime('%B %Y').upper(), 4)
                r += 1
                current_month = mkey
            ws.cell(r, 1, d).number_format = DATE_FMT
            ws.cell(r, 2, float(gmv or 0)).number_format = DOLLAR
            ws.cell(r, 3, orders or 0); ws.cell(r, 4, items or 0)
            t_gmv += gmv or Decimal('0'); t_orders += orders or 0; t_items += items or 0
            r += 1
        ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
        c = ws.cell(r, 2, float(t_gmv)); c.number_format = DOLLAR; c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.cell(r, 3, t_orders).font = F_TOTAL; ws.cell(r, 4, t_items).font = F_TOTAL
        ws.freeze_panes = 'A2'
        created.add('Source — Shop Analytics')

    # --- Source — Ad Spend ---
    ad_qs = AdSpendDay.objects.filter(date__gte=start_date, date__lte=end_date).order_by('date')
    ad_list = list(ad_qs.values_list('date', 'cost', 'sku_orders', 'gross_revenue'))
    if ad_list:
        ws = wb.create_sheet('Source — Ad Spend')
        _write_header(ws, ['Date', 'Cost', 'SKU Orders', 'Gross Revenue'], [12, 14, 12, 14])
        r = 2; current_month = None
        t_cost = Decimal('0'); t_orders = 0; t_rev = Decimal('0')
        for d, cost, orders, rev in ad_list:
            mkey = (d.year, d.month)
            if mkey != current_month:
                _section_header(ws, r, d.strftime('%B %Y').upper(), 4)
                r += 1
                current_month = mkey
            ws.cell(r, 1, d).number_format = DATE_FMT
            ws.cell(r, 2, float(cost or 0)).number_format = DOLLAR
            ws.cell(r, 3, orders or 0)
            ws.cell(r, 4, float(rev or 0)).number_format = DOLLAR
            t_cost += cost or Decimal('0'); t_orders += orders or 0; t_rev += rev or Decimal('0')
            r += 1
        ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
        c = ws.cell(r, 2, float(t_cost)); c.number_format = DOLLAR; c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.cell(r, 3, t_orders).font = F_TOTAL
        c = ws.cell(r, 4, float(t_rev)); c.number_format = DOLLAR; c.font = F_TOTAL; c.fill = FILL_TOTAL
        ws.freeze_panes = 'A2'
        created.add('Source — Ad Spend')

    # --- Source — Ad Ledger (only if FIFO engine ON) ---
    cfg = AdLedgerConfig.objects.filter(pk=1).first()
    if cfg and cfg.feed_pnl:
        al_qs = AdLedgerDay.objects.filter(date__gte=start_date, date__lte=end_date).order_by('date')
        al_list = list(al_qs.values_list('date', 'ad_spend', 'savings_tbsm', 'savings_promo'))
        if al_list:
            ws = wb.create_sheet('Source — Ad Ledger')
            _write_header(ws, ['Date', 'Spend', 'TBSM Savings', 'TT Promo Credits'], [12, 14, 14, 16])
            r = 2; current_month = None
            t_spend = Decimal('0'); t_tbsm = Decimal('0'); t_promo = Decimal('0')
            for d, spend, tbsm, promo in al_list:
                mkey = (d.year, d.month)
                if mkey != current_month:
                    _section_header(ws, r, d.strftime('%B %Y').upper(), 4)
                    r += 1
                    current_month = mkey
                ws.cell(r, 1, d).number_format = DATE_FMT
                ws.cell(r, 2, float(spend or 0)).number_format = DOLLAR
                ws.cell(r, 3, float(tbsm or 0)).number_format = DOLLAR
                ws.cell(r, 4, float(promo or 0)).number_format = DOLLAR
                t_spend += spend or Decimal('0'); t_tbsm += tbsm or Decimal('0'); t_promo += promo or Decimal('0')
                r += 1
            ws.cell(r, 1, 'GRAND TOTAL').font = F_TOTAL
            for col, val in zip([2, 3, 4], [t_spend, t_tbsm, t_promo]):
                c = ws.cell(r, col, float(val)); c.number_format = DOLLAR
                c.font = F_TOTAL; c.fill = FILL_TOTAL
            ws.freeze_panes = 'A2'
            created.add('Source — Ad Ledger')

    return created, skipped_for_size


def export_pnl(request):
    """GET: show the export page. With ?mode= & ?month= query: stream XLSX.
    Monthly + Multi-month exports include a Line Item Guide and Source — *
    sheets containing the raw imported rows for each line item, filtered to
    the export's date range. Click any line label in Sheet 1 → land in the
    Line Item Guide; click the guide row's "View Source Data" → land in the
    raw source rows for that line."""
    import io
    mode = request.GET.get('mode')
    yyyy_mm = request.GET.get('month', '')
    from_mm = request.GET.get('from_month', '')
    to_mm = request.GET.get('to_month', '')

    if not mode:
        return render(request, 'core/export.html', {
            'current_month': date.today().strftime('%Y-%m'),
        })
    if mode not in ('monthly', 'daily'):
        return HttpResponse('mode must be monthly or daily', status=400)

    # Multi-month range — monthly mode only, both from/to set
    is_range = (mode == 'monthly' and from_mm and to_mm)
    if is_range:
        try:
            fy, fm = int(from_mm[:4]), int(from_mm[5:7])
            ty, tm = int(to_mm[:4]), int(to_mm[5:7])
            for yr, mo in [(fy, fm), (ty, tm)]:
                if yr < 2020 or yr > 2099 or mo < 1 or mo > 12:
                    return HttpResponse('Invalid month range', status=400)
            if (fy, fm) > (ty, tm):
                return HttpResponse('from_month must be on or before to_month', status=400)
        except Exception:
            return HttpResponse('Invalid month range', status=400)
    else:
        try:
            y, m = int(yyyy_mm[:4]), int(yyyy_mm[5:7])
            if y < 2020 or y > 2099 or m < 1 or m > 12:
                return HttpResponse('Invalid month', status=400)
        except Exception:
            return HttpResponse('Invalid month', status=400)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from .line_item_docs import LINE_ITEM_DOCS

    wb = openpyxl.Workbook()
    ws = wb.active

    # Style presets
    F_HEAD = Font(bold=True, color='FFFFFF', size=11)
    F_SECTION = Font(bold=True, color='FFFFFF', size=11)
    F_SUB = Font(bold=True, size=10)
    F_TOTAL = Font(bold=True, size=11)
    F_LINK = Font(color='0563C1', underline='single', size=10)
    F_MONTH = Font(bold=True, size=12, color='FFFFFF')
    FILL_HEAD = PatternFill('solid', fgColor='2D3748')
    FILL_SECTION = PatternFill('solid', fgColor='4A5568')
    FILL_SUB = PatternFill('solid', fgColor='E2E8F0')
    FILL_TOTAL = PatternFill('solid', fgColor='EDF2F7')
    FILL_MONTH = PatternFill('solid', fgColor='4A5568')
    THIN = Side(style='thin', color='CBD5E0')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    SOURCE_STYLES = {
        'head': F_HEAD, 'fill_head': FILL_HEAD,
        'total': F_TOTAL, 'fill_total': FILL_TOTAL,
        'month': F_MONTH, 'fill_month': FILL_MONTH,
        'border': BORDER,
    }

    # Date range for source sheets — set per mode below
    src_start = src_end = None
    if mode == 'monthly' and is_range:
        src_start = date(fy, fm, 1)
        src_end = date(ty, tm, monthrange(ty, tm)[1])
    elif mode == 'monthly':
        src_start = date(y, m, 1)
        src_end = date(y, m, monthrange(y, m)[1])
    # Daily mode: skip source sheets (Daily P&L already shows the breakdown)

    # ---- Build source sheets FIRST (need their names for the Guide hyperlinks) ----
    sheets_present = set()
    sheets_skipped_size = set()
    if src_start and src_end:
        sheets_present, sheets_skipped_size = _build_source_sheets(
            wb, src_start, src_end, SOURCE_STYLES)

    # ---- Build 'Line Item Guide' (with View Source Data column) ----
    ws_doc = wb.create_sheet('Line Item Guide')
    ws_doc['A1'] = 'Line Item'
    ws_doc['B1'] = 'What it is'
    ws_doc['C1'] = 'Source'
    ws_doc['D1'] = 'Formula'
    ws_doc['E1'] = 'Notes'
    ws_doc['F1'] = 'View Source Data'
    for col, w_ in zip('ABCDEF', [38, 60, 50, 60, 70, 28]):
        ws_doc.column_dimensions[col].width = w_
    for c in ws_doc[1]:
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = Alignment(vertical='top')
    doc_row_for_label = {}
    doc_r = 2
    for label, doc in LINE_ITEM_DOCS.items():
        ws_doc.cell(doc_r, 1, label).font = F_TOTAL
        ws_doc.cell(doc_r, 2, doc.get('what', '')).alignment = Alignment(wrap_text=True, vertical='top')
        ws_doc.cell(doc_r, 3, doc.get('source', '')).alignment = Alignment(wrap_text=True, vertical='top')
        ws_doc.cell(doc_r, 4, doc.get('formula', '')).alignment = Alignment(wrap_text=True, vertical='top')
        ws_doc.cell(doc_r, 5, doc.get('notes', '')).alignment = Alignment(wrap_text=True, vertical='top')
        src_sheet = LINE_ITEM_TO_SOURCE.get(label.strip()) or LINE_ITEM_TO_SOURCE.get(label)
        if src_sheet and src_sheet in sheets_present:
            link_cell = ws_doc.cell(doc_r, 6, f'→ {src_sheet}')
            link_cell.alignment = Alignment(wrap_text=True, vertical='top')
            _set_internal_link(link_cell, src_sheet, 'A1', link_font=F_LINK)
        elif src_sheet and src_sheet in sheets_skipped_size:
            ws_doc.cell(doc_r, 6,
                'Range too large — export each month individually for raw rows '
                '(or use a 1-2 month range for full source data).'
            ).alignment = Alignment(wrap_text=True, vertical='top')
        else:
            ws_doc.cell(doc_r, 6, '— (no source export)').alignment = Alignment(wrap_text=True, vertical='top')
        ws_doc.row_dimensions[doc_r].height = 48
        for c in ws_doc[doc_r]: c.border = BORDER
        doc_row_for_label[label.strip()] = doc_r
        doc_r += 1
    ws_doc.freeze_panes = 'A2'

    def link_label_cell(cell, label):
        """Make P&L line label a working hyperlink → Line Item Guide row.
        Uses explicit Hyperlink object so Google Sheets honors it (string form
        was being ignored — that was the 'click goes nowhere' bug)."""
        target_row = doc_row_for_label.get(label.strip())
        if target_row:
            _set_internal_link(cell, 'Line Item Guide', f'A{target_row}', link_font=F_LINK)

    # ---- Sheet 1 — the P&L ----
    if mode == 'monthly' and is_range:
        # Multi-month range: one $/%NR column-pair per month + a Total range pair.
        months = []
        cy, cm = fy, fm
        while (cy, cm) <= (ty, tm):
            months.append((cy, cm))
            cm += 1
            if cm > 12:
                cm = 1; cy += 1

        monthly_per = {}
        nr_per = {}
        for (yr, mo) in months:
            start = date(yr, mo, 1)
            end = date(yr, mo, monthrange(yr, mo)[1])
            daily = compute_daily_pnl(start, end)
            m_totals = {}
            for d, row in daily.items():
                for label, val in row.items():
                    m_totals[label] = m_totals.get(label, Decimal('0')) + (val or Decimal('0'))
            monthly_per[(yr, mo)] = m_totals
            nr_per[(yr, mo)] = m_totals.get('NET REVENUE') or Decimal('0')
        range_nr = sum((nr_per[k] for k in months), Decimal('0'))

        ws.title = f'Monthly P&L {from_mm}..{to_mm}'

        # Two-row header: month labels (merged across $/%NR) on row 1, $/%NR on row 2
        hdr1 = ['Line Item']
        for yr, mo in months:
            hdr1.extend([date(yr, mo, 1).strftime('%b %Y'), ''])
        hdr1.extend([f'Total {from_mm}–{to_mm}', ''])
        ws.append(hdr1)
        hdr2 = ['']
        for _ in months: hdr2.extend(['$', '% NR'])
        hdr2.extend(['$', '% NR'])
        ws.append(hdr2)
        for c in ws[1]:
            c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = Alignment(horizontal='center')
        for c in ws[2]:
            c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = Alignment(horizontal='center')
        for i, _ in enumerate(months):
            col = 2 + i * 2
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        total_col = 2 + len(months) * 2
        ws.merge_cells(start_row=1, start_column=total_col, end_row=1, end_column=total_col + 1)
        # Uniform per-month widths: every month's $ col = 16, every %NR col = 11.
        # Same widths apply to the trailing Total range pair so all months look identical.
        ws.column_dimensions['A'].width = 42
        for i in range(2, len(hdr1) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16 if (i % 2 == 0) else 11

        excel_row = 3
        for label, rtype in PNL_ROW_LAYOUT:
            if rtype == 'blank':
                ws.row_dimensions[excel_row].height = 8
                excel_row += 1; continue
            if rtype in ('section', 'sub'):
                cell = ws.cell(excel_row, 1, label)
                if rtype == 'section':
                    cell.font = F_SECTION; cell.fill = FILL_SECTION
                else:
                    cell.font = F_SUB; cell.fill = FILL_SUB
                excel_row += 1; continue
            label_cell = ws.cell(excel_row, 1, label)
            link_label_cell(label_cell, label)
            if rtype == 'total':
                label_cell.font = F_TOTAL
            col = 2
            range_total = Decimal('0')
            for m_key in months:
                v = monthly_per[m_key].get(label)
                nr_m = nr_per[m_key]
                if v is not None:
                    c1 = ws.cell(excel_row, col, float(v))
                    c1.number_format = '$#,##0.00;[Red]($#,##0.00)'
                    if nr_m and float(nr_m) != 0:
                        c2 = ws.cell(excel_row, col + 1, float(v) / float(nr_m))
                        c2.number_format = '0.0%;[Red](0.0%)'
                    range_total += v
                    if rtype == 'total':
                        c1.font = F_TOTAL; c1.fill = FILL_TOTAL
                col += 2
            c1 = ws.cell(excel_row, col, float(range_total))
            c1.number_format = '$#,##0.00;[Red]($#,##0.00)'
            if range_nr and float(range_nr) != 0:
                c2 = ws.cell(excel_row, col + 1, float(range_total) / float(range_nr))
                c2.number_format = '0.0%;[Red](0.0%)'
                if rtype == 'total':
                    c2.font = F_TOTAL; c2.fill = FILL_TOTAL
            if rtype == 'total':
                c1.font = F_TOTAL; c1.fill = FILL_TOTAL
            excel_row += 1
        ws.freeze_panes = 'B3'
        filename = f'pnl_monthly_{from_mm}_to_{to_mm}.xlsx'

    elif mode == 'monthly':
        ws.title = f'Monthly P&L {yyyy_mm}'
        start = date(y, m, 1)
        end = date(y, m, monthrange(y, m)[1])
        daily = compute_daily_pnl(start, end)
        monthly = {}
        for d, row in daily.items():
            for label, val in row.items():
                monthly[label] = monthly.get(label, Decimal('0')) + (val or Decimal('0'))
        nr = monthly.get('NET REVENUE') or Decimal('0')

        # Header
        ws.append(['Line Item', f'{yyyy_mm} ($)', f'{yyyy_mm} (% Net Rev)'])
        for c in ws[1]:
            c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 16

        excel_row = 2
        for label, rtype in PNL_ROW_LAYOUT:
            if rtype == 'blank':
                ws.row_dimensions[excel_row].height = 8
                excel_row += 1; continue
            if rtype in ('section', 'sub'):
                cell = ws.cell(excel_row, 1, label)
                if rtype == 'section':
                    cell.font = F_SECTION; cell.fill = FILL_SECTION
                else:
                    cell.font = F_SUB; cell.fill = FILL_SUB
                excel_row += 1; continue
            v = monthly.get(label)
            label_cell = ws.cell(excel_row, 1, label)
            link_label_cell(label_cell, label)
            if rtype == 'total':
                label_cell.font = F_TOTAL
                for c in [ws.cell(excel_row, 2), ws.cell(excel_row, 3)]:
                    c.font = F_TOTAL; c.fill = FILL_TOTAL
            if v is not None:
                val_cell = ws.cell(excel_row, 2, float(v))
                val_cell.number_format = '$#,##0.00;[Red]($#,##0.00)'
                if nr and float(nr) != 0:
                    pct_cell = ws.cell(excel_row, 3, float(v) / float(nr))
                    pct_cell.number_format = '0.0%;[Red](0.0%)'
            excel_row += 1
        ws.freeze_panes = 'B2'
        filename = f'pnl_monthly_{yyyy_mm}.xlsx'

    else:  # daily
        ws.title = f'Daily P&L {yyyy_mm}'
        start = date(y, m, 1)
        end = date(y, m, monthrange(y, m)[1])
        daily = compute_daily_pnl(start, end)
        dates = sorted(daily.keys())
        nr_per_date = [daily.get(d, {}).get('NET REVENUE') for d in dates]
        month_nr = sum((v or Decimal('0')) for v in nr_per_date)

        # Two-row header: dates + $/% sub-header
        hdr1 = ['Line Item']
        for d in dates:
            hdr1.extend([d.strftime('%d %b'), ''])
        hdr1.extend([f'Total {yyyy_mm}', ''])
        ws.append(hdr1)
        hdr2 = ['']
        for _ in dates: hdr2.extend(['$', '% NR'])
        hdr2.extend(['$', '% NR'])
        ws.append(hdr2)
        for c in ws[1]:
            c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = Alignment(horizontal='center')
        for c in ws[2]:
            c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 42
        for i in range(2, len(hdr1) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 11

        excel_row = 3
        for label, rtype in PNL_ROW_LAYOUT:
            if rtype == 'blank':
                ws.row_dimensions[excel_row].height = 8
                excel_row += 1; continue
            if rtype in ('section', 'sub'):
                cell = ws.cell(excel_row, 1, label)
                if rtype == 'section':
                    cell.font = F_SECTION; cell.fill = FILL_SECTION
                else:
                    cell.font = F_SUB; cell.fill = FILL_SUB
                excel_row += 1; continue
            label_cell = ws.cell(excel_row, 1, label)
            link_label_cell(label_cell, label)
            if rtype == 'total':
                label_cell.font = F_TOTAL
            col = 2
            month_total = Decimal('0')
            for d, nr in zip(dates, nr_per_date):
                v = daily.get(d, {}).get(label)
                if v is not None:
                    c1 = ws.cell(excel_row, col, float(v))
                    c1.number_format = '$#,##0;[Red]($#,##0)'
                    if nr and float(nr) != 0:
                        c2 = ws.cell(excel_row, col + 1, float(v) / float(nr))
                        c2.number_format = '0.0%;[Red](0.0%)'
                    month_total += v
                col += 2
            c1 = ws.cell(excel_row, col, float(month_total))
            c1.number_format = '$#,##0;[Red]($#,##0)'
            if rtype == 'total':
                c1.font = F_TOTAL; c1.fill = FILL_TOTAL
            if month_nr and float(month_nr) != 0:
                c2 = ws.cell(excel_row, col + 1, float(month_total) / float(month_nr))
                c2.number_format = '0.0%;[Red](0.0%)'
                if rtype == 'total':
                    c2.font = F_TOTAL; c2.fill = FILL_TOTAL
            excel_row += 1
        ws.freeze_panes = 'B3'
        filename = f'pnl_daily_{yyyy_mm}.xlsx'

    # Stream out
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def ad_discounts(request):
    """Daily TikTok ad-account ledger: spend, balance, funding source, effective discount.
    Manages opening balance, agency-promo tags, and agency invoices."""
    from datetime import datetime as _dt
    from .ad_ledger import recompute_ledger

    def _parse_date(s):
        if not s: return None
        try: return _dt.strptime(s, '%Y-%m-%d').date()
        except Exception: return None

    try:
        year = int(request.GET.get('year', 2026))
        if year < 2020 or year > 2099: year = 2026
    except (TypeError, ValueError):
        year = 2026

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'save_config':
            cfg, _c = AdLedgerConfig.objects.get_or_create(pk=1)
            cfg.opening_balance = Decimal(request.POST.get('opening_balance', '0') or '0')
            cfg.opening_date = _parse_date(request.POST.get('opening_date'))
            cfg.opening_discount = Decimal(request.POST.get('opening_discount', '0.06') or '0.06')
            cfg.tbsm_default_discount = Decimal(request.POST.get('tbsm_default_discount', '0.06') or '0.06')
            cfg.feed_pnl = (request.POST.get('feed_pnl') == 'on')
            cfg.save()
            recompute_ledger(date(year, 1, 1), date(year, 12, 31))
            messages.success(request, 'Config saved + ledger recomputed.')
        elif action == 'add_tag':
            d = _parse_date(request.POST.get('tag_date'))
            if d:
                AgencyPromoTag.objects.update_or_create(
                    date=d,
                    defaults={
                        'min_amount': Decimal(request.POST.get('tag_min', '10000') or '10000'),
                        'discount_pct': Decimal(request.POST.get('tag_disc', '0.10') or '0.10'),
                        'note': request.POST.get('tag_note', ''),
                    },
                )
                recompute_ledger(date(year, 1, 1), date(year, 12, 31))
                messages.success(request, f'Agency promo tag saved for {d}.')
        elif action == 'delete_tag':
            try:
                AgencyPromoTag.objects.filter(pk=int(request.POST.get('tag_id'))).delete()
                recompute_ledger(date(year, 1, 1), date(year, 12, 31))
                messages.success(request, 'Tag deleted.')
            except (TypeError, ValueError):
                pass
        elif action == 'add_invoice':
            AgencyInvoice.objects.create(
                invoice_no=request.POST.get('inv_no', ''),
                issue_date=_parse_date(request.POST.get('inv_date')),
                loaded_value=Decimal(request.POST.get('inv_loaded', '0') or '0'),
                discount_pct=Decimal(request.POST.get('inv_disc', '0.06') or '0.06'),
                amount_paid=Decimal(request.POST.get('inv_paid', '0') or '0'),
                entity=request.POST.get('inv_entity', ''),
                notes=request.POST.get('inv_notes', ''),
            )
            messages.success(request, 'Invoice added.')
        elif action == 'delete_invoice':
            try:
                AgencyInvoice.objects.filter(pk=int(request.POST.get('inv_id'))).delete()
                messages.success(request, 'Invoice deleted.')
            except (TypeError, ValueError):
                pass
        elif action == 'recompute':
            recompute_ledger(date(year, 1, 1), date(year, 12, 31))
            messages.success(request, f'Ledger recomputed for {year}.')
        return redirect(f"{request.path}?year={year}")

    cfg, _c = AdLedgerConfig.objects.get_or_create(pk=1)
    days = list(AdLedgerDay.objects.filter(date__year=year).order_by('date'))
    tags = AgencyPromoTag.objects.all()
    invoices = AgencyInvoice.objects.all()
    txn_count = AdTransaction.objects.count()

    # Monthly rollup
    monthly = {}
    for d in days:
        m = d.date.strftime('%Y-%m')
        if m not in monthly:
            monthly[m] = {'spend': Decimal('0'), 'funded': Decimal('0'),
                          'full_price': Decimal('0'), 'card': Decimal('0'),
                          'actual_cost': Decimal('0'),
                          'savings_tbsm': Decimal('0'), 'savings_promo': Decimal('0'),
                          'end_balance': Decimal('0')}
        m_row = monthly[m]
        m_row['spend'] += d.ad_spend
        m_row['funded'] += d.funded
        m_row['full_price'] += d.full_price
        m_row['card'] += d.card_charge
        m_row['actual_cost'] += d.actual_cost
        m_row['savings_tbsm'] += d.savings_tbsm
        m_row['savings_promo'] += d.savings_promo
        m_row['end_balance'] = d.closing_balance
    for m_key, row in monthly.items():
        row['eff_disc'] = (
            (Decimal('1') - row['actual_cost'] / row['spend']) * 100
            if row['spend'] else Decimal('0'))
    monthly_sorted = sorted(monthly.items())

    return render(request, 'core/ad_discounts.html', {
        'year': year, 'prev_year': year - 1, 'next_year': year + 1,
        'days': days, 'monthly': monthly_sorted,
        'config': cfg, 'tags': tags, 'invoices': invoices,
        'txn_count': txn_count,
    })


def health(request):
    return HttpResponse('OK')


@csrf_exempt
def wipe(request):
    """Wipe imported data (keeps COGS + Monthly Inputs). POST only — protected by app password middleware."""
    if request.method != 'POST':
        return HttpResponse('Use POST. Optionally ?what=orders|settlement|analytics|ad_spend|all', status=405)
    what = request.GET.get('what', 'all')
    counts = {}
    if what in ('all', 'orders'):
        counts['orders'] = __import__('core.models', fromlist=['Order']).Order.objects.all().delete()[0]
    if what in ('all', 'settlement'):
        counts['settlement'] = __import__('core.models', fromlist=['SettlementRow']).SettlementRow.objects.all().delete()[0]
    if what in ('all', 'analytics'):
        counts['analytics'] = __import__('core.models', fromlist=['AnalyticsDay']).AnalyticsDay.objects.all().delete()[0]
    if what in ('all', 'ad_spend'):
        counts['ad_spend'] = __import__('core.models', fromlist=['AdSpendDay']).AdSpendDay.objects.all().delete()[0]
    if what in ('all', 'seller_shipping'):
        counts['seller_shipping'] = __import__('core.models', fromlist=['SellerShipmentCost']).SellerShipmentCost.objects.all().delete()[0]
    if what in ('all', 'monthly_inputs'):
        counts['monthly_inputs'] = __import__('core.models', fromlist=['MonthlyInput']).MonthlyInput.objects.all().delete()[0]
        counts['monthly_audit'] = __import__('core.models', fromlist=['MonthlyInputAudit']).MonthlyInputAudit.objects.all().delete()[0]
    if what in ('all', 'logs'):
        counts['logs'] = __import__('core.models', fromlist=['ImportLog']).ImportLog.objects.all().delete()[0]
    return HttpResponse('Wiped: ' + str(counts))
