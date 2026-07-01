"""
File parsers for the 5 TikTok exports.
All dates handled as Python date objects internally — no Sheet locale ambiguity.
"""
import csv
import io
from datetime import date, datetime
from decimal import Decimal
from django.db import transaction
import openpyxl

from .models import (Order, SettlementRow, AnalyticsDay, AdSpendDay, MonthlyInput,
                     COGSItem, ImportLog, SellerShipmentCost, AdTransaction,
                     FBTBillingSchedule)


def _to_dec(v):
    if v is None or v == '': return Decimal('0')
    if isinstance(v, (int, float, Decimal)): return Decimal(str(v))
    s = str(v).replace(',', '').replace('$', '').replace('\t', '').strip()
    s = s.replace('(', '-').replace(')', '')
    try: return Decimal(s)
    except: return Decimal('0')


def _to_int(v):
    if v is None or v == '': return 0
    try: return int(float(str(v).replace('\t', '').strip()))
    except: return 0


def _to_date(v, prefer_dd_mm=False):
    """Parse a date. Handles ambiguous MM/DD vs DD/MM by checking validity.
    `prefer_dd_mm=True` for files known to use DD/MM/YYYY (Shop Analytics, Ad Spend)."""
    if not v: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip().replace('\t', '').split(' ')[0]
    # Unambiguous YYYY-first
    for fmt in ('%Y/%m/%d', '%Y-%m-%d'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    # Ambiguous slash-separated — disambiguate via numeric checks
    parts = s.split('/')
    if len(parts) == 3:
        try:
            p1, p2, p3 = int(parts[0]), int(parts[1]), int(parts[2])
            # Year must be 4 digits and = p3
            if p3 < 100: p3 += 2000
            if p1 > 12 and p2 <= 12:    # MUST be DD/MM/YYYY
                return date(p3, p2, p1)
            if p2 > 12 and p1 <= 12:    # MUST be MM/DD/YYYY
                return date(p3, p1, p2)
            if p1 <= 12 and p2 <= 12:   # AMBIGUOUS — use hint
                if prefer_dd_mm:
                    return date(p3, p2, p1)
                else:
                    return date(p3, p1, p2)
        except (ValueError, TypeError):
            pass
    return None


def _clean_str(v):
    if v is None: return ''
    return str(v).strip().replace('\t', '')


def _cogs_lookup():
    return {c.sku_id: c.cogs_per_order for c in COGSItem.objects.all()}


# ===========================================================================
# 1. Manage Orders CSV importer
# ===========================================================================
def import_manage_orders(file_obj, filename=''):
    """Memory-efficient: streams CSV, bulk-inserts in 2000-row chunks, discards each chunk."""
    # Read as text but use iter line-by-line to keep memory bounded
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(raw))
    header = next(reader, None)
    del raw  # free
    if not header:
        return {'added': 0, 'skipped': 0, 'errors': ['Empty file']}

    hdr = [h.strip().lower() for h in header]
    def col(name):
        try: return hdr.index(name.lower())
        except ValueError: return -1

    c_oid = col('Order ID')
    c_status = col('Order Status')
    c_sku = col('SKU ID')
    c_qty = col('Quantity')
    c_gross = col('SKU Subtotal Before Discount')
    c_disc = col('SKU Seller Discount')
    c_created = col('Created Time')

    if c_oid < 0 or c_sku < 0 or c_created < 0:
        return {'added': 0, 'skipped': 0, 'errors': ['Missing required columns']}

    cogs_map = _cogs_lookup()
    # Fetch existing rows AND their current status so we can update if status changed
    existing_status = {(o['order_id'], o['sku_id']): o['status']
                       for o in Order.objects.values('order_id', 'sku_id', 'status')}

    chunk = []
    CHUNK_SIZE = 2000
    added_total = 0
    skipped = 0
    unmapped = set()
    status_changes = []  # (order_id, sku_id, new_status) for re-imports where status updated

    def flush(c):
        if not c: return 0
        with transaction.atomic():
            Order.objects.bulk_create(c, batch_size=500, ignore_conflicts=True)
        return len(c)

    for row in reader:
        if len(row) < max(c_oid, c_sku, c_created) + 1: continue
        oid = _clean_str(row[c_oid])
        sku = _clean_str(row[c_sku])
        if not oid or not sku: continue

        new_status = _clean_str(row[c_status] if c_status >= 0 and c_status < len(row) else '')

        if (oid, sku) in existing_status:
            # Already imported. If status changed (e.g., Shipped → Canceled),
            # update so the aggregator's Canceled-exclude filter applies.
            if existing_status[(oid, sku)] != new_status and new_status:
                status_changes.append((oid, sku, new_status))
            skipped += 1
            continue
        existing_status[(oid, sku)] = new_status

        # Manage Orders export uses MM/DD/YYYY (e.g., "05/31/2026 8:59:36 PM").
        # Explicit prefer_dd_mm=False so the default is documented + future-proof.
        created = _to_date(row[c_created] if c_created < len(row) else None, prefer_dd_mm=False)
        if not created: continue
        qty = _to_int(row[c_qty] if c_qty < len(row) else 0)
        gross = _to_dec(row[c_gross] if c_gross >= 0 and c_gross < len(row) else 0)
        disc = -abs(_to_dec(row[c_disc] if c_disc >= 0 and c_disc < len(row) else 0))
        cogs_per = cogs_map.get(sku)
        if cogs_per is None:
            unmapped.add(sku); cogs_val = Decimal('0')
        else:
            cogs_val = Decimal(cogs_per) * qty

        chunk.append(Order(
            order_id=oid, sku_id=sku, created_date=created, quantity=qty,
            status=new_status, gross_sale=gross, seller_discount=disc,
            cogs=cogs_val, source_file=filename,
        ))
        if len(chunk) >= CHUNK_SIZE:
            added_total += flush(chunk)
            chunk = []

    added_total += flush(chunk)

    # Apply status updates for re-imports (handles late cancellations)
    status_updated = 0
    if status_changes:
        with transaction.atomic():
            for oid, sku, new_status in status_changes:
                Order.objects.filter(order_id=oid, sku_id=sku).update(status=new_status)
                status_updated += 1

    notes_parts = []
    if unmapped: notes_parts.append('Unmapped SKUs: ' + ','.join(sorted(unmapped)))
    if status_updated: notes_parts.append(f'Status updated on {status_updated} existing orders')
    ImportLog.objects.create(importer='manage_orders', filename=filename,
                             rows_added=added_total, rows_skipped=skipped,
                             notes=' | '.join(notes_parts))
    return {'added': added_total, 'skipped': skipped,
            'status_updated': status_updated,
            'unmapped_skus': sorted(unmapped)}


# ===========================================================================
# 2. Settlement xlsx importer
# ===========================================================================
TYPE_TO_FIELD = {
    'Logistics reimbursement': 'logistics_reimb',
    'FBT warehouse service fee using GMV payment': 'fbt_warehouse',
    'Chargeback': 'chargeback',
    'Violation fee （settlement fee）': 'violation',
    'Violation fee (settlement fee)': 'violation',
    'TikTok Shop reimbursement': 'tt_shop_reimb',
    'FBT warehouse compensation': 'fbt_warehouse_comp',
    'Rebate': 'rebate',
}

def import_settlement(file_obj, filename=''):
    # Use pandas instead of openpyxl — read_only mode has a row-iteration bug,
    # and full openpyxl load uses 500MB+ which OOMs even on Standard tier.
    # pandas uses ~175MB for the same file. Read all columns as text to avoid type coercion.
    import pandas as pd
    df = pd.read_excel(file_obj, sheet_name=0, engine='openpyxl', dtype=str, keep_default_na=False)
    header = list(df.columns)
    h_lower = [str(h).strip().lower() for h in header]
    def col(name):
        try: return h_lower.index(name.lower())
        except ValueError: return -1

    c_stmt = col('Statement date')
    c_stid = col('Statement ID')
    c_type = col('Type')
    c_oid = col('Order/adjustment ID')
    c_qty = col('Quantity')
    c_created = col('Order created date')
    c_gross_ref = col('Gross sales refund')
    c_disc_ref = col('Seller discount refund')
    c_shipping = col('Shipping')
    c_tt_inc = col('TikTok Shop shipping incentive')
    c_tt_inc_refund = col('TikTok Shop shipping incentive refund')
    c_subsidy = col('Shipping fee subsidy')
    c_offset = col('Customer shipping fee offset')
    c_cust_paid = col('Customer-paid shipping fee')
    c_cust_paid_refund = col('Customer-paid shipping fee refund')
    c_fbt_fee = col('FBT fulfillment fee')
    c_fbt_reimb = col('FBT fulfillment fee reimbursement')
    c_ref = col('Referral fee')
    c_refadmin = col('Refund administration fee')
    c_aff = col('Affiliate Commission')
    c_aff_p = col('Affiliate partner commission')
    c_aff_sa = col('Affiliate Shop Ads commission')
    c_aff_psa = col('Affiliate Partner shop ads commission')
    c_cof = col('Co-funded promotion (seller-funded)')
    c_cof_camp = col('Co-funded Promotion campaign period fee')
    c_seller_ship_disc = col('Seller shipping fee discount')
    c_camp = col('Campaign service fee')
    c_adj = col('Adjustment amount')

    if c_oid < 0 or c_type < 0 or c_stid < 0:
        return {'added': 0, 'skipped': 0, 'errors': ['Missing required Settlement columns']}

    existing = set(SettlementRow.objects.values_list('order_id', 'settlement_id', 'row_type'))
    CHUNK_SIZE = 2000
    added_total = 0
    skipped = 0
    duplicate_merges = 0
    unknown_types = {}

    # Fields the upsert path will refresh on a pre-existing row. ALL numeric
    # fields go here so a re-upload of a fresher settlement file overwrites
    # stale values (previously this list was shipping-only, which caused
    # affiliate_total + every other fee field to silently NOT update on
    # re-import — e.g. May affiliate stuck at an old value).
    UPSERT_FIELDS = [
        'referral_fee', 'affiliate_total', 'campaign_fee', 'refund_admin',
        'fbt_fee', 'fbt_reimb', 'shipping', 'tt_ship_net',
        'tt_shop_shipping_incentive', 'tt_shop_shipping_incentive_refund',
        'shipping_fee_subsidy',
        'customer_shipping_fee_offset', 'customer_paid_shipping_fee',
        'customer_paid_shipping_refund',
        'cofunded_promo', 'cofunded_promo_campaign_fee',
        'seller_shipping_fee_discount', 'refund_total',
        'chargeback', 'violation', 'tt_shop_reimb', 'logistics_reimb',
        'fbt_warehouse', 'fbt_warehouse_comp', 'rebate', 'unclassified',
        'quantity',
    ]

    # Numeric fields to sum when the same key appears more than once in the file.
    # Multi-SKU orders produce one settlement row per SKU sharing the same
    # (order_id, statement_id, row_type) — they need to be collapsed to one
    # order-level row by summing every dollar/quantity field.
    SUM_FIELDS = [
        'quantity', 'referral_fee', 'affiliate_total', 'campaign_fee',
        'refund_admin', 'fbt_fee', 'fbt_reimb', 'shipping',
        'tt_shop_shipping_incentive', 'tt_shop_shipping_incentive_refund',
        'shipping_fee_subsidy',
        'customer_shipping_fee_offset', 'customer_paid_shipping_fee',
        'customer_paid_shipping_refund', 'tt_ship_net',
        'cofunded_promo', 'cofunded_promo_campaign_fee',
        'seller_shipping_fee_discount', 'refund_total',
        'chargeback', 'violation', 'tt_shop_reimb', 'logistics_reimb',
        'fbt_warehouse', 'fbt_warehouse_comp', 'rebate', 'unclassified',
    ]

    def flush_settle(c):
        if not c: return 0
        with transaction.atomic():
            SettlementRow.objects.bulk_create(
                c, batch_size=500,
                update_conflicts=True,
                unique_fields=['order_id', 'settlement_id', 'row_type'],
                update_fields=UPSERT_FIELDS,
            )
        return len(c)

    # ---- Pass 1: read every row, deduplicate by (oid, sid, rt) ----
    sr_by_key = {}
    for row in df.itertuples(index=False, name=None):
        if not row or not row[c_type]: continue
        oid = _clean_str(row[c_oid])
        sid = _clean_str(row[c_stid])
        rt = _clean_str(row[c_type])
        if not oid or not sid or not rt: continue
        key = (oid, sid, rt)

        created = _to_date(row[c_created]) if c_created >= 0 else None
        stmt = _to_date(row[c_stmt]) if c_stmt >= 0 else None
        date_for_log = created or stmt

        sr = SettlementRow(
            order_created_date=date_for_log,
            statement_date=stmt,
            order_id=oid, settlement_id=sid, row_type=rt,
            quantity=_to_int(row[c_qty]) if c_qty >= 0 else 0,
            source_file=filename,
        )

        if rt == 'Order':
            sr.referral_fee = _to_dec(row[c_ref]) if c_ref >= 0 else 0
            sr.affiliate_total = (
                _to_dec(row[c_aff]) if c_aff >= 0 else 0) + (
                _to_dec(row[c_aff_p]) if c_aff_p >= 0 else 0) + (
                _to_dec(row[c_aff_sa]) if c_aff_sa >= 0 else 0) + (
                _to_dec(row[c_aff_psa]) if c_aff_psa >= 0 else 0)
            sr.campaign_fee = _to_dec(row[c_camp]) if c_camp >= 0 else 0
            sr.refund_admin = _to_dec(row[c_refadmin]) if c_refadmin >= 0 else 0
            sr.fbt_fee = _to_dec(row[c_fbt_fee]) if c_fbt_fee >= 0 else 0
            sr.fbt_reimb = _to_dec(row[c_fbt_reimb]) if c_fbt_reimb >= 0 else 0
            sr.shipping = _to_dec(row[c_shipping]) if c_shipping >= 0 else 0
            tt_inc = _to_dec(row[c_tt_inc]) if c_tt_inc >= 0 else 0
            ship_sub = _to_dec(row[c_subsidy]) if c_subsidy >= 0 else 0
            cust_off = _to_dec(row[c_offset]) if c_offset >= 0 else 0
            sr.tt_shop_shipping_incentive = tt_inc
            sr.tt_shop_shipping_incentive_refund = _to_dec(row[c_tt_inc_refund]) if c_tt_inc_refund >= 0 else 0
            sr.shipping_fee_subsidy = ship_sub
            sr.customer_shipping_fee_offset = cust_off
            sr.customer_paid_shipping_fee = _to_dec(row[c_cust_paid]) if c_cust_paid >= 0 else 0
            sr.customer_paid_shipping_refund = _to_dec(row[c_cust_paid_refund]) if c_cust_paid_refund >= 0 else 0
            sr.tt_ship_net = tt_inc + ship_sub + cust_off
            sr.cofunded_promo = _to_dec(row[c_cof]) if c_cof >= 0 else 0
            sr.cofunded_promo_campaign_fee = _to_dec(row[c_cof_camp]) if c_cof_camp >= 0 else 0
            sr.seller_shipping_fee_discount = _to_dec(row[c_seller_ship_disc]) if c_seller_ship_disc >= 0 else 0
            sr.refund_total = (
                _to_dec(row[c_gross_ref]) if c_gross_ref >= 0 else 0) + (
                _to_dec(row[c_disc_ref]) if c_disc_ref >= 0 else 0)
        else:
            adj = _to_dec(row[c_adj]) if c_adj >= 0 else 0
            field = TYPE_TO_FIELD.get(rt)
            if field:
                setattr(sr, field, adj)
            else:
                sr.unclassified = adj
                unknown_types[rt] = unknown_types.get(rt, 0) + 1

        # Merge duplicates: same key already seen → sum numeric fields onto the
        # existing entry; keep the first non-numeric values.
        if key in sr_by_key:
            prior = sr_by_key[key]
            for f in SUM_FIELDS:
                cur = getattr(prior, f, None)
                add = getattr(sr, f, None)
                if cur is None: cur = 0
                if add is None: add = 0
                setattr(prior, f, cur + add)
            duplicate_merges += 1
        else:
            sr_by_key[key] = sr

    # ---- Pass 2: flush deduped rows in chunks ----
    chunk = []
    for sr in sr_by_key.values():
        key = (sr.order_id, sr.settlement_id, sr.row_type)
        if key in existing:
            skipped += 1
        chunk.append(sr)
        if len(chunk) >= CHUNK_SIZE:
            added_total += flush_settle(chunk)
            chunk = []
    added_total += flush_settle(chunk)

    # Reports tab drift detector — compare per-order sums to the Reports tab
    # parent values from THIS file. Catches: (a) TikTok adding/renaming columns
    # we don't read, (b) importer bugs that drop columns silently.
    drift_msgs = _verify_reports_tab_drift(file_obj, df, c_type)

    note_parts = []
    if duplicate_merges:
        note_parts.append(f'Multi-SKU rows merged: {duplicate_merges:,}')
    if unknown_types:
        note_parts.append('Unknown types: ' + '; '.join(f'{k}={v}' for k, v in unknown_types.items()))
    if drift_msgs:
        note_parts.append('Drift vs Reports tab: ' + '; '.join(drift_msgs))
    note = ' | '.join(note_parts)
    ImportLog.objects.create(importer='settlement', filename=filename,
                             rows_added=added_total, rows_skipped=skipped, notes=note)
    return {'added': added_total, 'skipped': skipped, 'unknown_types': unknown_types,
            'drift': drift_msgs}


def _verify_reports_tab_drift(file_obj, df, c_type):
    """Read the Reports tab from a settlement file, compare its parent values to
    per-order sums from the Order details sheet. Returns a list of drift messages
    (empty if everything matches within $1)."""
    import pandas as pd
    # Per-order sums from the df we already loaded (Order rows only)
    od = df[df[df.columns[c_type]].astype(str).str.strip() == 'Order'] if c_type >= 0 else df
    cols_lower = {c.lower().strip(): c for c in df.columns}
    def col_by_name(name):
        return cols_lower.get(name.lower())
    def sum_col(name):
        c = col_by_name(name)
        if c is None: return None
        return float(pd.to_numeric(od[c], errors='coerce').fillna(0).sum())

    # Read Reports tab
    try:
        file_obj.seek(0)
    except Exception:
        return ['Could not re-read file for drift check']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as e:
        return [f'Could not open file for Reports tab: {e}']
    if 'Reports' not in wb.sheetnames:
        return ['Reports tab not found in file']
    ws = wb['Reports']
    reports = {}
    for row in ws.iter_rows(values_only=True):
        if not row: continue
        cells = [c for c in row if c is not None and str(c).strip() != '']
        if len(cells) >= 2:
            label = str(cells[-2]).strip()
            try:
                reports[label] = float(cells[-1])
            except (TypeError, ValueError):
                pass

    # Compare the parent columns we should match
    checks = [
        'Gross sales', 'Gross sales refund', 'Seller discount', 'Seller discount refund',
        'Shipping', 'TikTok Shop shipping incentive', 'Shipping fee subsidy',
        'Customer shipping fee offset', 'Customer-paid shipping fee',
        'Customer-paid shipping fee refund', 'FBT fulfillment fee',
        'FBT fulfillment fee reimbursement', 'Referral fee',
        'Refund administration fee', 'Affiliate Commission',
        'Affiliate partner commission', 'Affiliate Shop Ads commission',
        'Affiliate Partner shop ads commission',
        'Co-funded promotion (seller-funded)',
        'Co-funded Promotion campaign period fee',
        'Campaign service fee', 'Seller shipping fee discount',
    ]
    drifts = []
    for label in checks:
        rv = reports.get(label)
        ours = sum_col(label)
        if rv is None or ours is None:
            continue
        if abs(ours - rv) > 1.0:
            drifts.append(f'{label}: ours=${ours:,.2f} vs Reports=${rv:,.2f} (Δ ${ours - rv:+,.2f})')
    return drifts


# ===========================================================================
# 3. Shop Analytics importer
# ===========================================================================
def import_shop_analytics(file_obj, filename=''):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    start = -1
    for i, r in enumerate(rows):
        if r and str(r[0] or '').lower() == 'date':
            start = i + 1; break
    if start < 0:
        return {'added': 0, 'skipped': 0, 'errors': ['Could not find "Date" header']}

    # TikTok Shop Analytics always exports DD/MM/YYYY. Default to that.
    # Only switch to MM/DD if a date in this file PROVES MM/DD (part 2 > 12).
    prefer_dd_mm = True
    for r in rows[start:]:
        if not r or not r[0]: continue
        s = str(r[0]).strip().split(' ')[0]
        parts = s.split('/')
        if len(parts) == 3:
            try:
                p1, p2 = int(parts[0]), int(parts[1])
                if p1 > 12 and p2 <= 12:
                    prefer_dd_mm = True; break  # confirmed DD/MM
                if p2 > 12 and p1 <= 12:
                    prefer_dd_mm = False; break  # confirmed MM/DD — rare
            except: pass

    count = 0
    for r in rows[start:]:
        d = _to_date(r[0] if len(r) > 0 else None, prefer_dd_mm=prefer_dd_mm)
        gmv = _to_dec(r[1] if len(r) > 1 else 0)
        if not d: continue
        orders = _to_int(r[2] if len(r) > 2 else 0)
        items_sold = _to_int(r[4] if len(r) > 4 else 0)
        AnalyticsDay.objects.update_or_create(
            date=d, defaults={'gmv': gmv, 'orders': orders, 'items_sold': items_sold})
        count += 1
    ImportLog.objects.create(importer='shop_analytics', filename=filename, rows_added=count)
    return {'added': count, 'format_detected': 'DD/MM/YYYY' if prefer_dd_mm else 'MM/DD/YYYY'}


# ===========================================================================
# 4. Ad Spend importer
# ===========================================================================
def import_ad_spend(file_obj, filename=''):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    start = -1
    for i, r in enumerate(rows):
        if r and 'by day' in str(r[0] or '').lower():
            start = i + 1; break
    if start < 0:
        return {'added': 0, 'skipped': 0, 'errors': ['Could not find "By Day" header']}

    # Pre-scan format
    prefer_dd_mm = False
    for r in rows[start:]:
        if not r or not r[0]: continue
        s = str(r[0]).strip().split(' ')[0]
        parts = s.split('/')
        if len(parts) == 3:
            try:
                if int(parts[0]) > 12: prefer_dd_mm = True; break
            except: pass

    count = 0
    for r in rows[start:]:
        d = _to_date(r[0] if len(r) > 0 else None, prefer_dd_mm=prefer_dd_mm)
        cost = _to_dec(r[1] if len(r) > 1 else 0)
        if not d: continue
        sku_orders = _to_int(r[2] if len(r) > 2 else 0)
        gross_rev = _to_dec(r[4] if len(r) > 4 else 0)
        AdSpendDay.objects.update_or_create(
            date=d, defaults={'cost': cost, 'sku_orders': sku_orders, 'gross_revenue': gross_rev})
        count += 1
    ImportLog.objects.create(importer='ad_spend', filename=filename, rows_added=count)
    return {'added': count, 'format_detected': 'DD/MM/YYYY' if prefer_dd_mm else 'MM/DD/YYYY'}


# ===========================================================================
# 5. FBT Billing importer
# ===========================================================================
FBT_LINE_TO_FIELD = {
    'Hub placement fee': 'fbt_hub_placement',
    'Storage fee': 'fbt_storage',
    'Inbound shipping fee': 'fbt_inbound_shipping',
    'Inbound incidents fee': 'fbt_inbound_incidents',
    'Inbound booking non-compliance fee': 'fbt_booking_noncomp',
    'Routing non-compliance fee': 'fbt_routing_noncomp',
    'Outbound appointment no-show fee': 'fbt_outbound_noshow',
    'Delayed response fee': 'fbt_delayed_response',
    'Disposal fee': 'fbt_disposal',
    'Return shipping fee': 'fbt_return_shipping',
    'Return to seller handling fee': 'fbt_return_seller_handling',
    'Inbound return operation fee': 'fbt_inbound_return_op',
}

# ===========================================================================
# 6. Seller Shipping (per-shipment postage) CSV importer
# ===========================================================================
def import_seller_shipping(file_obj, filename=''):
    """Seller-shipping CSV importer. Supports two formats:

    LEGACY: channel_name, shipped_date, reference_number, shipment_number,
            carrier_service, tracking, postage, ...

    NEW (cost breakdown + order_date — preferred):
            order_date, shipped_date, reference_number, customer_name,
            postage, product_quantity, per_pack, per_pick

    In the new format, `reference_number` is the de-facto unique key (each row is
    a 3PL shipment line). It's stored in `shipment_number` so existing dedup logic
    keeps working. Order date drives the P&L attribution (accrual-consistent).
    Dates accept 'Apr 1, 2026', '2026-04-01', or '#N/A' (treated as null)."""
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(raw))
    header = next(reader, None)
    del raw
    if not header:
        return {'added': 0, 'skipped': 0, 'errors': ['Empty file']}

    hdr = [h.strip().lower() for h in header]
    def col(name):
        try: return hdr.index(name.lower())
        except ValueError: return -1

    c_chan = col('channel_name')
    c_order_date = col('order_date')
    c_ship_date = col('shipped_date')
    c_ref = col('reference_number')
    c_ship_num = col('shipment_number')
    c_customer = col('customer_name')
    c_carrier = col('carrier_service')
    c_track = col('tracking')
    c_post = col('postage')
    c_qty = col('product_quantity')
    c_per_pack = col('per_pack')
    c_per_pick = col('per_pick')

    # Determine which format. New format = has order_date + per_pack columns.
    is_new_format = c_order_date >= 0 and c_per_pack >= 0
    # The key column for dedup: use shipment_number if present (legacy), else reference_number (new).
    c_key = c_ship_num if c_ship_num >= 0 else c_ref
    if c_key < 0 or c_post < 0 or c_ship_date < 0:
        return {'added': 0, 'skipped': 0,
                'errors': ['Missing required columns (need shipment_number OR reference_number, plus shipped_date and postage)']}

    def parse_loose_date(s):
        """Parse 'Apr 1, 2026', '2026-04-01', etc. Returns None for '#N/A' or empty."""
        if not s or s.lower() in ('#n/a', 'na', 'n/a'): return None
        s = s.strip()
        for fmt in ('%b %d, %Y', '%B %d, %Y', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y'):
            try: return datetime.strptime(s, fmt).date()
            except Exception: pass
        return _to_date(s)

    # Materialize rows so we can pass over them twice: once for pre-clean,
    # once for actual import.
    all_rows = list(reader)

    # Pre-clean: if this file has REAL shipment_numbers (distinct from
    # reference_numbers on the same row), any existing DB rows keyed by
    # reference_number (from an older upload that didn't have the shipment_number
    # column) are the SAME physical shipments — leaving them in place would
    # double-count. Delete those old-format duplicates before insert.
    if c_ship_num >= 0 and c_ref >= 0 and c_ship_num != c_ref:
        new_refs_with_distinct_ship_num = set()
        for row in all_rows:
            if len(row) <= max(c_ship_num, c_ref): continue
            sn = _clean_str(row[c_ship_num])
            rn = _clean_str(row[c_ref])
            if sn and rn and sn != rn:
                new_refs_with_distinct_ship_num.add(rn)
        if new_refs_with_distinct_ship_num:
            # Old-format entries have shipment_number == reference_number.
            # Delete those matching this file's refs.
            SellerShipmentCost.objects.filter(
                shipment_number__in=new_refs_with_distinct_ship_num,
                reference_number__in=new_refs_with_distinct_ship_num,
            ).extra(where=['shipment_number = reference_number']).delete()

    existing = set(SellerShipmentCost.objects.values_list('shipment_number', flat=True))
    chunk = []
    CHUNK_SIZE = 2000
    added_total = 0
    skipped = 0

    UPDATE_FIELDS = [
        'order_date', 'shipped_date', 'postage', 'per_pack', 'per_pick',
        'product_quantity', 'reference_number', 'customer_name',
        'carrier_service', 'tracking', 'channel_name', 'source_file',
    ]

    def flush(c):
        if not c: return 0
        with transaction.atomic():
            SellerShipmentCost.objects.bulk_create(
                c, batch_size=500,
                update_conflicts=True,
                unique_fields=['shipment_number'],
                update_fields=UPDATE_FIELDS,
            )
        return len(c)

    for row in all_rows:
        if len(row) <= c_key: continue
        key = _clean_str(row[c_key])
        if not key: continue
        if key in existing:
            skipped += 1
        else:
            existing.add(key)

        d_ship = parse_loose_date(_clean_str(row[c_ship_date])
                                  if c_ship_date >= 0 and c_ship_date < len(row) else '')
        d_order = parse_loose_date(_clean_str(row[c_order_date])
                                    if c_order_date >= 0 and c_order_date < len(row) else '')
        if not d_ship and not d_order: continue
        # Shipped date is required for backward compat; if missing fall back to order date.
        if not d_ship: d_ship = d_order

        chunk.append(SellerShipmentCost(
            shipment_number=key,
            order_date=d_order,
            shipped_date=d_ship,
            postage=_to_dec(row[c_post]) if c_post < len(row) else Decimal('0'),
            per_pack=_to_dec(row[c_per_pack]) if c_per_pack >= 0 and c_per_pack < len(row) else Decimal('0'),
            per_pick=_to_dec(row[c_per_pick]) if c_per_pick >= 0 and c_per_pick < len(row) else Decimal('0'),
            product_quantity=_to_int(row[c_qty]) if c_qty >= 0 and c_qty < len(row) else 0,
            reference_number=_clean_str(row[c_ref]) if c_ref >= 0 and c_ref < len(row) else '',
            customer_name=_clean_str(row[c_customer]) if c_customer >= 0 and c_customer < len(row) else '',
            carrier_service=_clean_str(row[c_carrier]) if c_carrier >= 0 and c_carrier < len(row) else '',
            tracking=_clean_str(row[c_track]) if c_track >= 0 and c_track < len(row) else '',
            channel_name=_clean_str(row[c_chan]) if c_chan >= 0 and c_chan < len(row) else '',
            source_file=filename,
        ))
        if len(chunk) >= CHUNK_SIZE:
            added_total += flush(chunk); chunk = []

    added_total += flush(chunk)

    # Legacy cleanup: when a NEW-format file is imported, also delete any leftover
    # legacy rows. Legacy rows were keyed by `shipment_number` (3PL-internal IDs
    # like "CS12186116") while new-format rows are keyed by `reference_number`
    # (TikTok order IDs like "577410…" or short order codes). Same physical
    # shipment, two different keys → no dedup → double-count on Cost to Ship to
    # Customer. Legacy rows have per_pack=0 AND per_pick=0 (those columns didn't
    # exist when they were imported). New rows always have per_pack > 0.
    legacy_deleted = 0
    if is_new_format:
        with transaction.atomic():
            legacy_deleted, _ = SellerShipmentCost.objects.filter(
                per_pack=0, per_pick=0,
            ).delete()

    note_parts = [f'Format: {"new (with order_date/per_pack/per_pick)" if is_new_format else "legacy"}']
    if legacy_deleted:
        note_parts.append(f'Cleaned up {legacy_deleted:,} legacy rows (postage-only)')
    ImportLog.objects.create(
        importer='seller_shipping', filename=filename,
        rows_added=added_total, rows_skipped=skipped,
        notes=' | '.join(note_parts),
    )
    return {'added': added_total, 'skipped': skipped,
            'format': 'new' if is_new_format else 'legacy',
            'legacy_cleaned': legacy_deleted}


def import_fbt_billing(file_obj, period, filename=''):
    """period must be 'YYYY-MM'."""
    if not period or len(period) != 7 or period[4] != '-':
        return {'errors': ['Invalid period — expected YYYY-MM']}
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mi, _ = MonthlyInput.objects.get_or_create(month=period)
    count = 0
    for r in ws.iter_rows(values_only=True):
        if not r: continue
        line = _clean_str(r[1]) if len(r) > 1 else ''
        amt = _to_dec(r[2]) if len(r) > 2 else 0
        field = FBT_LINE_TO_FIELD.get(line)
        if field:
            setattr(mi, field, abs(amt))
            count += 1
    mi.save()
    ImportLog.objects.create(importer='fbt_billing', filename=filename,
                             rows_added=count, notes=f'Period: {period}')
    return {'added': count, 'period': period}


# ===========================================================================
# 6b. FBT Payment Cycle importer (TikTok Seller Center → Finance → Payment cycle)
# ===========================================================================
# Period names in the file look like "April 2026", "December 2025", etc.
_MONTH_NAMES = {
    'january': '01', 'february': '02', 'march': '03', 'april': '04',
    'may': '05', 'june': '06', 'july': '07', 'august': '08',
    'september': '09', 'october': '10', 'november': '11', 'december': '12',
}


def _parse_period(s):
    """Convert 'April 2026' → '2026-04'. Returns None on parse failure."""
    if not s: return None
    parts = str(s).strip().lower().split()
    if len(parts) != 2: return None
    month_num = _MONTH_NAMES.get(parts[0])
    if not month_num: return None
    try:
        year = int(parts[1])
        return f'{year:04d}-{month_num}'
    except ValueError:
        return None


def import_fbt_payment_cycle(file_obj, filename=''):
    """TikTok Payment Cycle XLSX → FBTBillingSchedule rows.

    The file has one row per (billing period, statement date) entry. We
    aggregate by summing the Settled amount across rows with the same
    (period, statement_date), then upsert. This gives the aggregator the
    statement_date for each services month so the 12 FBT-detail lines can
    attribute to the correct destination month (flat-spread within it).

    Expected columns: Billing period | Statement date | Due date | Total |
                       Settled | Outstanding | Status | Subject
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {'added': 0, 'errors': ['Empty file']}

    # Find columns from header row
    hdr = [_clean_str(c).lower() if c is not None else '' for c in rows[0]]
    def col(name):
        for i, h in enumerate(hdr):
            if h == name.lower(): return i
        return -1
    c_period = col('billing period')
    c_stmt = col('statement date')
    c_settled = col('settled')
    c_status = col('status')
    if c_period < 0 or c_stmt < 0 or c_settled < 0:
        return {'added': 0, 'errors': [
            f'Missing required columns. Found headers: {hdr}']}

    # Aggregate (period, stmt_date) → (total_settled, status_str)
    agg = {}
    for r in rows[1:]:
        if not r: continue
        period_raw = r[c_period] if c_period < len(r) else None
        stmt_raw = r[c_stmt] if c_stmt < len(r) else None
        settled_raw = r[c_settled] if c_settled < len(r) else None
        status = _clean_str(r[c_status]) if c_status >= 0 and c_status < len(r) else ''
        period = _parse_period(period_raw)
        stmt_date = _to_date(stmt_raw)
        amt = _to_dec(settled_raw)
        if not period or not stmt_date:
            continue
        key = (period, stmt_date)
        if key in agg:
            agg[key]['amount'] += amt
        else:
            agg[key] = {'amount': amt, 'status': status}

    # Upsert: existing rows for the same (period, stmt_date) get overwritten,
    # new rows are created. We don't delete missing rows — a re-uploaded file
    # might cover a narrower date range than the prior one.
    added = updated = 0
    for (period, stmt_date), data in agg.items():
        obj, was_created = FBTBillingSchedule.objects.update_or_create(
            period=period,
            statement_date=stmt_date,
            defaults={
                'amount': data['amount'],
                'status': data['status'],
                'source_file': filename,
            },
        )
        if was_created: added += 1
        else: updated += 1

    ImportLog.objects.create(
        importer='fbt_payment_cycle', filename=filename,
        rows_added=added,
        notes=f'Schedules: {len(agg)} ({added} new, {updated} updated)')
    return {'added': added, 'updated': updated, 'total': len(agg)}


# ===========================================================================
# 7. Ad Transactions importer (TikTok Ads Manager → Transactions XLSX)
# ===========================================================================
def import_ad_transactions(file_obj, filename=''):
    """Read 3 sheets (Payments / Promotions / Others) and upsert AdTransaction rows.
    Recomputes the AdLedgerDay snapshot over the file's date range."""
    import pandas as pd
    try:
        xl = pd.ExcelFile(file_obj)
    except Exception as e:
        return {'added': 0, 'skipped': 0, 'errors': [f'Could not open file: {e}']}

    required = {'Payments', 'Promotions', 'Others'}
    actual = set(xl.sheet_names)
    if not required.issubset(actual):
        return {'added': 0, 'skipped': 0,
                'errors': [f'Expected sheets {required}; got {actual}']}

    buf = []
    min_d = None
    max_d = None
    for sheet in ['Payments', 'Promotions', 'Others']:
        df = pd.read_excel(xl, sheet)
        # Normalize column lookup
        cols = {c.strip(): c for c in df.columns}
        def col(name):
            for k in cols:
                if k.lower() == name.lower(): return cols[k]
            return None
        c_time = col('Transaction time')
        c_id = col('Transaction ID')
        c_amt = col('Amount')
        c_type = col('Transaction type')
        c_status = col('Status')
        c_details = col('Details')
        c_typelabel = col('Type')

        if not c_time or not c_amt:
            continue

        for _, row in df.iterrows():
            txn_time = pd.to_datetime(row.get(c_time), errors='coerce')
            if pd.isna(txn_time): continue
            txn_id = str(row.get(c_id, '') or '').strip() if c_id else ''
            amount = _to_dec(row.get(c_amt))
            txn_type = (str(row.get(c_type, '') or '').strip() if c_type else '')
            status = (str(row.get(c_status, '') or '').strip() if c_status else '')
            details = (str(row.get(c_details, '') or '').strip() if c_details else '')
            type_label = (str(row.get(c_typelabel, '') or '').strip() if c_typelabel else '')
            buf.append(AdTransaction(
                txn_id=txn_id, txn_time=txn_time.to_pydatetime(), sheet=sheet,
                txn_type=txn_type, status=status, amount=amount,
                details=details, type_label=type_label, source_file=filename,
            ))
            d = txn_time.date()
            if min_d is None or d < min_d: min_d = d
            if max_d is None or d > max_d: max_d = d

    if not buf:
        return {'added': 0, 'skipped': 0, 'errors': ['No usable rows in file']}

    with transaction.atomic():
        AdTransaction.objects.bulk_create(
            buf, batch_size=500,
            update_conflicts=True,
            unique_fields=['txn_id', 'sheet', 'txn_time', 'amount'],
            update_fields=['txn_type', 'status', 'details', 'type_label', 'source_file'],
        )

    # Recompute the ledger over a slightly wider range so spend days outside the txn
    # window but inside the spend table also get updated balances.
    from .ad_ledger import recompute_ledger
    if min_d and max_d:
        # Extend to end-of-month on both sides for safety
        end_recompute = max_d.replace(day=1)
        from calendar import monthrange
        end_recompute = end_recompute.replace(day=monthrange(end_recompute.year, end_recompute.month)[1])
        recompute_ledger(min_d.replace(day=1), end_recompute)

    ImportLog.objects.create(importer='ad_transactions', filename=filename,
                             rows_added=len(buf), rows_skipped=0,
                             notes=f'Date range: {min_d} → {max_d}')
    return {'added': len(buf), 'skipped': 0, 'date_range': f'{min_d} → {max_d}'}
